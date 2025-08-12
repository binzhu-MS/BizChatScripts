#!/usr/bin/env python3
"""
Excel to TSV Converter

This script converts Excel files (.xlsx, .xls) to TSV (Tab-Separated Values) format
with UTF-8 encoding. It can handle multiple sheets and provides options for
customizing the conversion process.

Usage:
    python excel_to_tsv.py --input_file=data.xlsx
    python excel_to_tsv.py --input_file=data.xlsx --output_file=output.tsv
    python excel_to_tsv.py --input_file=data.xlsx --all_sheets=True
    python excel_to_tsv.py --input_file=data.xlsx --sheet_name="Sheet Name"

Requirements:
    pip install pandas openpyxl xlrd fire
"""

import os
import sys
from pathlib import Path
from typing import Optional, List, Union
import pandas as pd
import xlrd  # For .xls files
import openpyxl  # For .xlsx files
import fire  # For command line interface
from fire.core import FireExit  # For handling Fire's exit exceptions


def get_excel_sheets(excel_file: str) -> List[str]:
    """
    Get list of sheet names from Excel file.

    Args:
        excel_file: Path to Excel file

    Returns:
        List of sheet names
    """
    try:
        print(f"Attempting to read Excel file: {excel_file}")
        print(f"File exists: {os.path.exists(excel_file)}")
        if os.path.exists(excel_file):
            file_size = os.path.getsize(excel_file)
            print(f"File size: {file_size:,} bytes")

            # Check file header to identify actual file type
            with open(excel_file, "rb") as f:
                header = f.read(512)  # Read more bytes for better analysis
                print(f"First 32 bytes (hex): {header[:32].hex()}")
                print(
                    f"First 32 bytes (ascii, errors ignored): {repr(header[:32].decode('ascii', errors='ignore'))}"
                )

                # Check for specific file signatures
                if header.startswith(b"PK"):
                    print("File appears to be a ZIP-based format (expected for .xlsx)")
                elif header.startswith(b"\xd0\xcf\x11\xe0"):
                    print("File appears to be in OLE2 format (older Excel .xls format)")
                    # Check if it contains Excel-specific signatures
                    if b"Microsoft Excel" in header or b"Excel" in header:
                        print("  - Contains Excel signature")
                    if b"Workbook" in header:
                        print("  - Contains Workbook signature")
                    if b"\x09\x08" in header:
                        print("  - Contains BIFF signature")
                elif header.startswith(b"\x09\x08"):
                    print("File might be in older Excel BIFF format")
                else:
                    print(
                        "File format not recognized - may be corrupted or not an Excel file"
                    )
                    print(
                        f"Full header analysis needed. First 64 bytes: {header[:64].hex()}"
                    )

                # Look for Excel-specific strings
                excel_indicators = [b"Excel", b"Workbook", b"Sheet", b"Microsoft"]
                found_indicators = [
                    indicator for indicator in excel_indicators if indicator in header
                ]
                if found_indicators:
                    print(f"Found Excel indicators: {found_indicators}")
                else:
                    print(
                        "No Excel indicators found in header - file may not be a valid Excel file"
                    )

        # Try different engines based on actual file format, not just extension
        file_path = Path(excel_file)
        extension = file_path.suffix.lower()

        # Determine the correct engine based on file header
        with open(excel_file, "rb") as f:
            header = f.read(4)

        if header.startswith(b"PK"):
            # ZIP-based format (xlsx, xlsm, xlsb)
            print("File is ZIP-based format, using openpyxl engine...")
            try:
                with pd.ExcelFile(excel_file, engine="openpyxl") as xls:
                    sheets = [str(sheet) for sheet in xls.sheet_names]
                    print(f"Successfully read {len(sheets)} sheet(s) with openpyxl")
                    return sheets
            except Exception as openpyxl_error:
                print(f"openpyxl failed: {openpyxl_error}")
                # Fall through to other methods
        elif header.startswith(b"\xd0\xcf\x11\xe0"):
            # OLE2 format (xls) - but might actually be xlsx saved as xls
            print("File is OLE2 format, trying multiple approaches...")

            # First try: pandas with xlrd engine
            try:
                print("Trying pandas with xlrd engine...")
                with pd.ExcelFile(excel_file, engine="xlrd") as xls:
                    sheets = [str(sheet) for sheet in xls.sheet_names]
                    print(f"Successfully read {len(sheets)} sheet(s) with pandas+xlrd")
                    return sheets
            except Exception as xlrd_error:
                print(f"pandas+xlrd failed: {xlrd_error}")

            # Second try: direct xlrd with ignore_workbook_corruption
            try:
                print("Trying direct xlrd with ignore_workbook_corruption...")
                import xlrd

                wb = xlrd.open_workbook(excel_file, ignore_workbook_corruption=True)
                sheets = [str(sheet) for sheet in wb.sheet_names()]
                print(
                    f"Direct xlrd with ignore_workbook_corruption succeeded: {len(sheets)} sheet(s)"
                )
                return sheets
            except Exception as xlrd_alt_error:
                print(
                    f"Direct xlrd with ignore_workbook_corruption failed: {xlrd_alt_error}"
                )

            # Third try: Try openpyxl anyway (sometimes OLE2 files can be read by openpyxl)
            try:
                print("Trying openpyxl on OLE2 file (sometimes works)...")
                with pd.ExcelFile(excel_file, engine="openpyxl") as xls:
                    sheets = [str(sheet) for sheet in xls.sheet_names]
                    print(
                        f"Successfully read {len(sheets)} sheet(s) with openpyxl on OLE2 file"
                    )
                    return sheets
            except Exception as openpyxl_error:
                print(f"openpyxl on OLE2 file failed: {openpyxl_error}")

            # Fourth try: Try with pandas default (no engine specified)
            try:
                print("Trying pandas with default engine...")
                with pd.ExcelFile(excel_file) as xls:
                    sheets = [str(sheet) for sheet in xls.sheet_names]
                    print(
                        f"Successfully read {len(sheets)} sheet(s) with pandas default engine"
                    )
                    return sheets
            except Exception as default_error:
                print(f"pandas default engine failed: {default_error}")

            print("All OLE2 reading methods failed")
        else:
            # Fall back to extension-based detection and try all engines
            print(f"Unknown file format, trying all available engines...")

        # Try all engines systematically if header-based detection didn't work
        engines_to_try = []
        file_path = Path(excel_file)
        extension = file_path.suffix.lower()

        if extension == ".xlsx":
            engines_to_try = ["openpyxl", "xlrd", None]  # None = pandas default
        elif extension == ".xls":
            engines_to_try = ["xlrd", "openpyxl", None]
        else:
            engines_to_try = ["openpyxl", "xlrd", None]

        for engine in engines_to_try:
            try:
                engine_name = engine if engine else "default"
                print(f"Trying pandas with {engine_name} engine...")
                with pd.ExcelFile(excel_file, engine=engine) as xls:
                    sheets = [str(sheet) for sheet in xls.sheet_names]
                    print(
                        f"Successfully read {len(sheets)} sheet(s) with {engine_name} engine"
                    )
                    return sheets
            except Exception as engine_error:
                print(f"{engine_name} engine failed: {engine_error}")
                continue

        # If all pandas approaches failed, try direct library access
        print("All pandas engines failed, trying direct library access...")

        # Try direct openpyxl
        try:
            print("Trying direct openpyxl...")
            import openpyxl

            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            sheets = [str(sheet) for sheet in wb.sheetnames]
            wb.close()
            print(f"Direct openpyxl succeeded: {len(sheets)} sheet(s)")
            return sheets
        except Exception as direct_openpyxl_error:
            print(f"Direct openpyxl failed: {direct_openpyxl_error}")

        # Try direct xlrd with various options
        try:
            print("Trying direct xlrd with various options...")
            import xlrd

            for ignore_corruption in [False, True]:
                for formatting_info in [False, True]:
                    try:
                        wb = xlrd.open_workbook(
                            excel_file,
                            ignore_workbook_corruption=ignore_corruption,
                            formatting_info=formatting_info,
                        )
                        sheets = [str(sheet) for sheet in wb.sheet_names()]
                        print(
                            f"Direct xlrd succeeded (ignore_corruption={ignore_corruption}, formatting_info={formatting_info}): {len(sheets)} sheet(s)"
                        )
                        return sheets
                    except Exception:
                        continue
        except Exception as direct_xlrd_error:
            print(f"Direct xlrd completely failed: {direct_xlrd_error}")

        # Try pyxlsb for Excel Binary Workbook format
        try:
            print("Trying pyxlsb for binary Excel format...")
            import pyxlsb

            with pyxlsb.open_workbook(excel_file) as wb:
                sheets = list(wb.sheets)
                print(f"pyxlsb succeeded: {len(sheets)} sheet(s)")
                return sheets
        except ImportError:
            print("pyxlsb not available")
        except Exception as pyxlsb_error:
            print(f"pyxlsb failed: {pyxlsb_error}")

        # Last resort: Use Windows COM to interact with Excel directly
        try:
            print("Trying Windows COM Excel automation (like opening in Excel)...")
            import win32com.client
            import pythoncom

            # Initialize COM
            pythoncom.CoInitialize()

            try:
                # Create Excel application with better error handling
                excel_app = win32com.client.DispatchEx("Excel.Application")
                excel_app.Visible = False  # Don't show Excel window
                excel_app.DisplayAlerts = False  # Don't show alerts

                try:
                    # Open the workbook (use absolute path)
                    full_path = os.path.abspath(excel_file)
                    print(f"Opening file with Excel COM: {full_path}")
                    workbook = excel_app.Workbooks.Open(full_path, ReadOnly=True)

                    # Get sheet names
                    sheets = []
                    for worksheet in workbook.Worksheets:
                        sheets.append(str(worksheet.Name))

                    # Close workbook
                    workbook.Close(SaveChanges=False)
                    print(f"Windows COM Excel succeeded: {len(sheets)} sheet(s)")
                    return sheets

                except Exception as wb_error:
                    print(f"Error opening workbook with COM: {wb_error}")

                finally:
                    # Always quit Excel application
                    try:
                        excel_app.Application.Quit()
                    except:
                        pass

            finally:
                # Uninitialize COM
                pythoncom.CoUninitialize()

        except Exception as com_error:
            print(f"Windows COM Excel failed: {com_error}")

        # Try with different file extensions to see if it's misnamed
        print("File might be misnamed, trying to read as different formats...")
        for try_extension in [".xlsb", ".xls", ".xlsx"]:
            try:
                print(f"Trying to read as {try_extension} format...")
                temp_path = Path(excel_file).with_suffix(try_extension)
                # Don't actually rename, just try different engines
                if try_extension == ".xlsb":
                    try:
                        import pyxlsb

                        with pyxlsb.open_workbook(excel_file) as wb:
                            sheets = list(wb.sheets)
                            print(f"Reading as .xlsb succeeded: {len(sheets)} sheet(s)")
                            return sheets
                    except Exception as xlsb_error:
                        print(f"Reading as .xlsb failed: {xlsb_error}")
                elif try_extension == ".xlsx":
                    try:
                        with pd.ExcelFile(excel_file, engine="openpyxl") as xls:
                            sheets = [str(sheet) for sheet in xls.sheet_names]
                            print(f"Reading as .xlsx succeeded: {len(sheets)} sheet(s)")
                            return sheets
                    except Exception:
                        pass
                elif try_extension == ".xls":
                    try:
                        with pd.ExcelFile(excel_file, engine="xlrd") as xls:
                            sheets = [str(sheet) for sheet in xls.sheet_names]
                            print(f"Reading as .xls succeeded: {len(sheets)} sheet(s)")
                            return sheets
                    except Exception:
                        pass
            except Exception as format_error:
                print(f"Format {try_extension} failed: {format_error}")

        print(
            "All Excel reading methods failed - file may be corrupted, password-protected, or in an unsupported format"
        )
        print("Suggestions:")
        print("1. Check if the file is password-protected")
        print("2. Try opening and re-saving the file in Excel")
        print("3. Check if the file is actually an Excel Binary Workbook (.xlsb)")
        return []

    except Exception as e:
        print(f"Error reading Excel file with primary method: {e}")
        print(f"Error type: {type(e).__name__}")

        # Try alternative approach with direct openpyxl/xlrd
        try:
            file_path = Path(excel_file)
            extension = file_path.suffix.lower()

            if extension == ".xlsx":
                print("Trying direct openpyxl approach...")
                import openpyxl

                wb = openpyxl.load_workbook(excel_file, read_only=True)
                sheets = [str(sheet) for sheet in wb.sheetnames]
                wb.close()
                print(f"Direct openpyxl succeeded: {len(sheets)} sheet(s)")
                return sheets

            elif extension == ".xls":
                print("Trying direct xlrd approach...")
                import xlrd

                wb = xlrd.open_workbook(excel_file)
                sheets = [str(sheet) for sheet in wb.sheet_names()]
                print(f"Direct xlrd succeeded: {len(sheets)} sheet(s)")
                return sheets

        except Exception as e2:
            print(f"Alternative approach also failed: {e2}")
            print(f"Alternative error type: {type(e2).__name__}")

        return []


def read_excel_sheet_with_com(excel_file: str, sheet_name: str) -> pd.DataFrame:
    """
    Read Excel sheet data using Windows COM (Excel automation).
    This works exactly like opening the file in Excel.

    Args:
        excel_file: Path to Excel file
        sheet_name: Name of the sheet to read

    Returns:
        pandas DataFrame with the sheet data
    """
    import win32com.client
    import pythoncom

    # Initialize COM
    pythoncom.CoInitialize()

    try:
        # Create Excel application
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False

        try:
            # Open workbook
            full_path = os.path.abspath(excel_file)
            workbook = excel_app.Workbooks.Open(full_path, ReadOnly=True)

            # Find the worksheet
            worksheet = None
            for ws in workbook.Worksheets:
                if ws.Name == sheet_name:
                    worksheet = ws
                    break

            if worksheet is None:
                raise ValueError(f"Sheet '{sheet_name}' not found")

            # Get the used range
            used_range = worksheet.UsedRange
            if used_range is None:
                # Empty sheet
                return pd.DataFrame()

            # Get data as 2D array
            data = used_range.Value

            # Close workbook
            workbook.Close(SaveChanges=False)

            # Convert to pandas DataFrame
            if data is None:
                return pd.DataFrame()

            # Handle single cell case
            if not isinstance(data, (list, tuple)):
                return pd.DataFrame([[data]])

            # Handle single row case
            if not isinstance(data[0], (list, tuple)):
                data = [data]

            # Create DataFrame
            df = pd.DataFrame(data)

            # Use first row as headers if it looks like headers
            if len(df) > 0 and df.iloc[0].notna().any():
                # Check if first row looks like headers (contains strings)
                first_row = df.iloc[0]
                if any(
                    isinstance(val, str) and val.strip()
                    for val in first_row
                    if val is not None
                ):
                    df.columns = [
                        str(col) if col is not None else f"Column_{i}"
                        for i, col in enumerate(first_row)
                    ]
                    df = df.drop(df.index[0]).reset_index(drop=True)

            return df

        finally:
            # Always quit Excel
            try:
                excel_app.Application.Quit()
            except:
                pass

    finally:
        # Uninitialize COM
        pythoncom.CoUninitialize()


def convert_excel_to_tsv(
    excel_file: str,
    output_file: Optional[str] = None,
    sheet_name: Optional[str] = None,
    all_sheets: bool = False,
    encoding: str = "utf-8",
    na_rep: str = "",
    index: bool = False,
    header: bool = True,
) -> bool:
    """
    Convert Excel file to TSV format.

    Args:
        excel_file: Path to input Excel file
        output_file: Path to output TSV file (optional)
        sheet_name: Specific sheet to convert (optional)
        all_sheets: Convert all sheets to separate TSV files
        encoding: Output file encoding (default: utf-8)
        na_rep: String representation of NaN values
        index: Whether to include row indices in output
        header: Whether to include column headers in output

    Returns:
        True if conversion successful, False otherwise
    """

    # Validate input file
    if not os.path.exists(excel_file):
        print(f"Error: Input file '{excel_file}' does not exist.")
        return False

    excel_path = Path(excel_file)
    if excel_path.suffix.lower() not in [".xlsx", ".xls"]:
        print(f"Error: Input file must be .xlsx or .xls format.")
        return False

    try:
        # Get sheet names
        sheet_names = get_excel_sheets(excel_file)
        if not sheet_names:
            print("Error: No sheets found in Excel file.")
            return False

        print(
            f"Found {len(sheet_names)} sheet(s): {', '.join(str(name) for name in sheet_names)}"
        )

        # Determine which sheets to convert
        sheets_to_convert = []

        if all_sheets:
            sheets_to_convert = sheet_names
        elif sheet_name:
            if sheet_name in sheet_names:
                sheets_to_convert = [sheet_name]
            else:
                print(
                    f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(str(name) for name in sheet_names)}"
                )
                return False
        else:
            # Default to first sheet
            sheets_to_convert = [sheet_names[0]]
            print(f"Converting first sheet: '{sheet_names[0]}'")

        # Convert each sheet
        for sheet in sheets_to_convert:
            print(f"Converting sheet: '{sheet}'...")

            # Read Excel sheet
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet)
                print(f"  Loaded {len(df)} rows, {len(df.columns)} columns")
            except Exception as e:
                print(f"  Error reading sheet '{sheet}' with pandas: {e}")
                # Try to read using Windows COM as fallback
                try:
                    print(f"  Trying Windows COM to read sheet '{sheet}'...")
                    df = read_excel_sheet_with_com(excel_file, sheet)
                    print(
                        f"  COM succeeded: Loaded {len(df)} rows, {len(df.columns)} columns"
                    )
                except Exception as com_e:
                    print(f"  Error reading sheet '{sheet}' with COM: {com_e}")
                    continue

            # Determine output filename
            if all_sheets:
                # Multiple files for multiple sheets - save in same directory as input file
                base_name = excel_path.stem
                input_dir = excel_path.parent
                safe_sheet_name = "".join(
                    c for c in str(sheet) if c.isalnum() or c in (" ", "-", "_")
                ).strip()
                tsv_file = input_dir / f"{base_name}_{safe_sheet_name}.tsv"
            elif output_file:
                tsv_file = output_file
            else:
                # Default output filename - save in same directory as input file
                tsv_file = excel_path.with_suffix(".tsv")

            # Convert to TSV
            try:
                df.to_csv(
                    tsv_file,
                    sep="\t",
                    encoding=encoding,
                    na_rep=na_rep,
                    index=index,
                    header=header,
                )
                print(f"  ✓ Saved to: {tsv_file}")

                # Show file info
                tsv_path = Path(tsv_file)
                file_size = tsv_path.stat().st_size
                print(f"    File size: {file_size:,} bytes")

            except Exception as e:
                print(f"  Error saving TSV file: {e}")
                return False

        print("✓ Conversion completed successfully!")
        return True

    except Exception as e:
        print(f"Error during conversion: {e}")
        return False


def main(
    input_file: str,
    output_file: Optional[str] = None,
    sheet_name: Optional[str] = None,
    all_sheets: bool = False,
    encoding: str = "utf-8",
    na_rep: str = "",
    include_index: bool = False,
    no_header: bool = False,
    list_sheets: bool = False,
) -> None:
    """
    Convert Excel files to TSV format with UTF-8 encoding.

    Args:
        input_file: Input Excel file (.xlsx or .xls) (required)
        output_file: Output TSV file (optional, default: same name as input with .tsv extension)
        sheet_name: Specific sheet name to convert (default: first sheet)
        all_sheets: Convert all sheets to separate TSV files
        encoding: Output file encoding (default: utf-8)
        na_rep: String representation of NaN values (default: empty string)
        include_index: Include row indices in output
        no_header: Do not include column headers in output
        list_sheets: List all sheets in the Excel file and exit

    Example usage:
        # Convert Excel to TSV
        python excel_to_tsv.py --input_file=data.xlsx

        # Convert with custom output file
        python excel_to_tsv.py --input_file=data.xlsx --output_file=output.tsv

        # Convert specific sheet
        python excel_to_tsv.py --input_file=data.xlsx --sheet_name="Sheet2"

        # Convert all sheets
        python excel_to_tsv.py --input_file=data.xlsx --all_sheets=True

        # List available sheets
        python excel_to_tsv.py --input_file=data.xlsx --list_sheets=True

    """
    try:
        # Validate arguments
        if all_sheets and output_file:
            print("Warning: all_sheets specified, output filename will be ignored")

        # List sheets option
        if list_sheets:
            sheets = get_excel_sheets(input_file)
            if sheets:
                print(f"Sheets in '{input_file}':")
                for i, sheet in enumerate(sheets, 1):
                    print(f"  {i}. {sheet}")
            else:
                print(f"No sheets found in '{input_file}'")
            return

        # Perform conversion
        success = convert_excel_to_tsv(
            excel_file=input_file,
            output_file=output_file,
            sheet_name=sheet_name,
            all_sheets=all_sheets,
            encoding=encoding,
            na_rep=na_rep,
            index=include_index,
            header=not no_header,
        )

        if not success:
            raise RuntimeError("Conversion failed")

        sys.exit(0)

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    # Use fire for command line arguments
    try:
        fire.Fire(main)
    except FireExit as e:
        # Handle Fire's exit (including --help) gracefully in debug mode
        # FireExit with code 0 means successful exit (like --help)
        # FireExit with non-zero code means error exit
        sys.exit(e.code)
